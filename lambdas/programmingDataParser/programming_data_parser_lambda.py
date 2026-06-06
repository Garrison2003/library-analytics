"""
Library Analytics – Programming Statistics Lambda (Enhanced with DynamoDB)

Parses programming statistics .xlsx files uploaded to S3 and:
1. Writes current data to S3 (for fast API responses)
2. Writes historical data to DynamoDB (for trend analysis)

Features:
  - Extracts monthly in-person attendance and program count data
  - Identifies department/branch by 3-letter code in filename
  - Returns data organized by month in fiscal year format
  - Stores historical data for trend analysis
  - Supports branch-level filtering via query parameter

Trigger:  S3 PUT on uploads/programming/*.xlsx
API:      GET /programming?branch={CODE}
          GET /programming/history?branch={CODE}&months=12
"""

import json
import logging
import os
import re
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import unquote_plus

import boto3
import openpyxl
import pdfplumber
from botocore.exceptions import ClientError
from decimal import Decimal

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client("s3")
dynamodb = boto3.resource("dynamodb")

# ── Constants ────────────────────────────────────────────────────────────────

# Branch code to full name mapping
BRANCH_CODE_MAP = {
    "ALW": "Allegra Westbrooks Regional",
    "CAR": "Carmel",
    "TEL": "Teen Loft",
    "CHS": "Charlotte",
    "COM": "Community",
    "COR": "Cornelius",
    "DAV": "Davidson",
    "EAS": "East",
    "HCG": "Hickory Grove",
    "IMG": "Imaginon",
    "INR": "Independence Regional",
    "LAC": "Library Admin Center",
    "MAI": "Main",
    "MAT": "Matthews",
    "MNH": "Mint Hill",
    "MOB": "Mobile Library",
    "MTI": "Mountain Island",
    "MYP": "Myers Park",
    "NCR": "North County Regional",
    "NOR": "Northlake",
    "PIN": "Pineville",
    "PLZ": "Plaza Midwood",
    "SBL": "South Boulevard",
    "SCR": "South County Regional",
    "SGC": "Sugar Creek",
    "SPA": "Spangler",
    "SPK": "SouthPark Regional",
    "STC": "Steele Creek",
    "UCR": "University City Regional",
    "WBL": "West Boulevard",
    "WES": "West",
}

PROCESSED_DIR = "processed/programming"

# Date format used in PDF report rows: "09-Apr-2026"
_DATE_RE = re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$")

# Reverse mapping: partial branch name in filename → branch code
BRANCH_NAME_TO_CODE = {
    "allegra": "ALW",
    "westbrooks": "ALW",
    "carmel": "CAR",
    "charlotte": "CHS",
    "community": "COM",
    "cornelius": "COR",
    "davidson": "DAV",
    "east": "EAS",
    "hickory": "HCG",
    "imaginon": "IMG",
    "independence": "INR",
    "admin": "LAC",
    "main": "MAI",
    "matthews": "MAT",
    "mint": "MNH",
    "mobile": "MOB",
    "mountain": "MTI",
    "myers": "MYP",
    "north county": "NCR",
    "northlake": "NOR",
    "pineville": "PIN",
    "plaza": "PLZ",
    "south boulevard": "SBL",
    "south county": "SCR",
    "sugar": "SGC",
    "spangler": "SPA",
    "teen loft": "TEL",
    "southpark": "SPK",
    "steele": "STC",
    "university": "UCR",
    "west boulevard": "WBL",
    "west": "WES",
}

# 3-letter month abbreviation → month number (as in PDF filter criteria)
MONTH_ABBR_TO_NUM = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

# ── Helpers ──────────────────────────────────────────────────────────────────

def _env(name: str, fallback: str = "") -> str:
    """Get environment variable or fallback."""
    return os.environ.get(name, fallback)


def _safe_int(val: Any) -> int:
    """Safely convert value to int, returning 0 on failure."""
    if val is None:
        return 0
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _extract_branch_code(filename: str) -> Optional[str]:
    """
    Extract 3-letter branch code from filename.
    Example: "IMG Monthly Stats FY26-28.xlsx" → "IMG"
    """
    match = re.match(r"^([A-Z]{3})\s+", filename)
    return match.group(1) if match else None


def _get_branch_name(code: str) -> str:
    """Get full branch name from code, or return code if not found."""
    return BRANCH_CODE_MAP.get(code, code)


def _date_to_year_month(date_obj) -> Optional[str]:
    """
    Convert date object to YYYY-MM format for DynamoDB sort key.
    Example: datetime(2024, 5, 15) → "2024-05"
    """
    try:
        if hasattr(date_obj, "month") and hasattr(date_obj, "year"):
            return f"{date_obj.year:04d}-{date_obj.month:02d}"
    except (AttributeError, TypeError):
        pass
    return None


def _month_year_to_display(month: int, year: int) -> str:
    """Convert month (1-12) and year (24, 25...) to MM/YY format."""
    return f"{month:02d}/{year:02d}"


# ── Excel Parsing ────────────────────────────────────────────────────────────

def _find_data_sheet(workbook) -> Optional[str]:
    """
    Find the appropriate data sheet in the workbook.
    Looks for "Program 23-26", "Program 23-27", or similar patterns.
    Falls back to "Teen Programs" or "Juv Programs".
    """
    for sheet_name in workbook.sheetnames:
        if "Program 23" in sheet_name or "Program 24" in sheet_name:
            return sheet_name
    for sheet_name in ["Teen Programs", "Juv Programs"]:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    return None


def parse_workbook(file_bytes: bytes) -> Dict[str, Any]:
    """
    Parse programming .xlsx and extract monthly data.

    Returns dict with:
      data: List of monthly records with attendance & program counts
      months: List of MM/YY labels
      attendance: List of attendance values
      programs: List of program counts
      year_month_data: List of tuples (year_month string, attendance, programs)
    """
    file_obj = BytesIO(file_bytes)
    wb = openpyxl.load_workbook(file_obj, data_only=True)

    data_sheet_name = _find_data_sheet(wb)
    if not data_sheet_name:
        logger.warning("No data sheet found in workbook. Available: %s", wb.sheetnames)
        wb.close()
        return {
            "data": [],
            "months": [],
            "attendance": [],
            "programs": [],
            "year_month_data": [],
        }

    ws = wb[data_sheet_name]
    logger.info("Parsing sheet: %s", data_sheet_name)

    result_data = []
    months_list = []
    attendance_list = []
    programs_list = []
    year_month_data = []  # For DynamoDB storage

    # Iterate through rows starting from row 2 (row 1 is headers)
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            # Skip header row
            continue

        if not row or not row[0]:
            # Skip empty rows
            continue

        # Get date from first column
        date_obj = row[0] if row else None
        if not date_obj:
            continue

        # Convert date to YYYY-MM format for DynamoDB
        year_month = _date_to_year_month(date_obj)
        if not year_month:
            continue

        # Also get MM/YY format for display
        try:
            if hasattr(date_obj, "month") and hasattr(date_obj, "year"):
                month = date_obj.month
                year = date_obj.year
                month_display = _month_year_to_display(month, year)
            else:
                continue
        except (AttributeError, TypeError):
            continue

        # Column mapping (0-based):
        # Col 0: Date
        # Col 1: TOTAL ATTENDANCE IN-PERSON
        # Col 2: TOTAL PROGRAMS
        # (Further columns: TOTAL VIRTUAL ATTENDANCE, increases, etc.)

        attendance = _safe_int(row[1]) if len(row) > 1 else 0
        programs = _safe_int(row[2]) if len(row) > 2 else 0
        virtual_attendance = _safe_int(row[3]) if len(row) > 3 else 0

        # Skip rows with no meaningful data
        if attendance == 0 and programs == 0:
            continue

        months_list.append(month_display)
        attendance_list.append(attendance)
        programs_list.append(programs)

        result_data.append({
            "month": month_display,
            "attendance": attendance,
            "programs": programs,
            "virtual_attendance": virtual_attendance,
            "date": f"{year_month}-01",
        })

        # Store for DynamoDB (with year_month key)
        year_month_data.append({
            "year_month": year_month,
            "attendance": attendance,
            "programs": programs,
            "virtual_attendance": virtual_attendance,
        })

    wb.close()

    logger.info("Parsed %d months of data", len(result_data))

    return {
        "data": result_data,
        "months": months_list,
        "attendance": attendance_list,
        "programs": programs_list,
        "year_month_data": year_month_data,
    }


# ── PDF Parsing ──────────────────────────────────────────────────────────────

def _extract_branch_code_from_name(filename: str) -> Optional[str]:
    """
    Extract branch code from a filename that contains the full/partial branch name.
    Example: "Spangler_In-House_Programs_April_2026.pdf" → "SPA"
    Checks longer keys first so "north county" matches before "north".
    """
    # Normalise separators so "north_county" matches the key "north county"
    name_lower = filename.lower().replace("_", " ").replace("-", " ")
    for name, code in sorted(BRANCH_NAME_TO_CODE.items(), key=lambda x: -len(x[0])):
        if name in name_lower:
            return code
    return None


def parse_pdf_report(file_bytes: bytes) -> Optional[Dict[str, Any]]:
    """
    Parse an In-House or Outreach Programs PDF report exported from the ILS.

    Extracts:
      - year_month  : "YYYY-MM" from "Filter Criteria Applied: Program Date: Apr 2026"
      - programs    : Grand Summary total # of programs
      - attendance  : Grand Summary total attendance
      - is_outreach : True if this is an Outreach report, False for In-House
      - sessions    : List of per-session dicts with facilitator, program_name,
                      program_date, num_programs, total_attendance, outreach_site

    Returns None on parse failure.
    """
    try:
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            full_text = ""
            all_tables: List[List] = []
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
                for tbl in (page.extract_tables() or []):
                    all_tables.extend(tbl)

        # Detect report type from title (first 500 chars)
        is_outreach = bool(re.search(r"\boutreach\b", full_text[:500], re.IGNORECASE))

        # Extract month/year: "Program Date: Apr 2026(Included)"
        month_match = re.search(
            r"Program Date:\s+([A-Za-z]+)\s+(\d{4})", full_text, re.IGNORECASE
        )
        if not month_match:
            logger.error("PDF: could not find month/year in filter criteria")
            return None

        month_abbr = month_match.group(1).upper()[:3]
        year = int(month_match.group(2))
        month_num = MONTH_ABBR_TO_NUM.get(month_abbr)
        if not month_num:
            logger.error("PDF: unknown month abbreviation '%s'", month_abbr)
            return None

        year_month = f"{year}-{month_num:02d}"

        # Column layout differs by report type:
        #   In-House : [Facilitator, ProgramName, Date, #Programs, Attendance]
        #   Outreach : [Facilitator, Site, ProgramName, Date, #Programs, Attendance]
        if is_outreach:
            date_col, num_col, att_col = 3, 4, 5
        else:
            date_col, num_col, att_col = 2, 3, 4

        def _c(row_data: List, i: int) -> str:
            """Return cell i as a normalised string (newlines collapsed, stripped)."""
            v = row_data[i] if len(row_data) > i else None
            return str(v).strip().replace("\n", " ") if v is not None else ""

        programs = 0
        attendance = 0
        sessions: List[Dict] = []
        current_facilitator: Optional[str] = None
        current_site: Optional[str] = None
        current_program: Optional[str] = None

        for row in all_tables:
            if not row:
                continue

            first = _c(row, 0)

            # Grand Summary row — capture totals, then move on
            if "grand summary" in first.lower():
                non_empty = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if len(non_empty) >= 3:
                    try:
                        programs = int(non_empty[-2].replace(",", ""))
                        attendance = int(non_empty[-1].replace(",", ""))
                    except (ValueError, IndexError):
                        pass
                continue

            # Skip column-header rows (appear at top of each page)
            if "primary facilitator" in first.lower():
                continue

            # Update carry-forward state for facilitator / site / program.
            # These span multiple rows via merged cells; pdfplumber returns
            # None for the continuation cells, so we only update on non-empty.
            if first:
                current_facilitator = first
            if is_outreach:
                v1, v2 = _c(row, 1), _c(row, 2)
                if v1:
                    current_site = v1
                if v2:
                    current_program = v2
            else:
                v1 = _c(row, 1)
                if v1:
                    current_program = v1

            # Only emit a session record for rows that carry an individual date.
            # Subtotal/total rows have an empty date column and are skipped here.
            date_val = _c(row, date_col)
            if not _DATE_RE.match(date_val):
                continue

            try:
                program_date = datetime.strptime(date_val, "%d-%b-%Y").strftime("%Y-%m-%d")
            except ValueError:
                continue

            if current_facilitator and current_program:
                sessions.append({
                    "primary_facilitator": current_facilitator,
                    "program_name": current_program,
                    "program_date": program_date,
                    "num_programs": _safe_int(_c(row, num_col)),
                    "total_attendance": _safe_int(_c(row, att_col)),
                    "outreach_site": current_site if is_outreach else None,
                })

        # Text fallback if table extraction missed the Grand Summary row
        if programs == 0 and attendance == 0:
            grand_match = re.search(
                r"Grand Summary[:\s]+(\d+)\s+(\d+)", full_text, re.IGNORECASE
            )
            if grand_match:
                programs = int(grand_match.group(1))
                attendance = int(grand_match.group(2))

        logger.info(
            "PDF parsed: %s | programs=%d | attendance=%d | outreach=%s | sessions=%d",
            year_month, programs, attendance, is_outreach, len(sessions),
        )

        return {
            "year_month": year_month,
            "programs": programs,
            "attendance": attendance,
            "is_outreach": is_outreach,
            "sessions": sessions,
        }

    except Exception as exc:
        logger.exception("Error parsing PDF report: %s", str(exc))
        return None


# ── S3 I/O ───────────────────────────────────────────────────────────────────

def read_s3(bucket: str, key: str) -> bytes:
    """Read file from S3."""
    return s3.get_object(Bucket=bucket, Key=key)["Body"].read()


def write_json_s3(bucket: str, key: str, payload: dict) -> None:
    """Write JSON payload to S3."""
    s3.put_object(
        Bucket=bucket,
        Key=key,
        Body=json.dumps(payload, indent=2),
        ContentType="application/json",
    )
    logger.info("Wrote s3://%s/%s", bucket, key)


# ── DynamoDB I/O ─────────────────────────────────────────────────────────────

def write_programming_data_to_dynamodb(
    branch_code: str,
    branch_name: str,
    year_month_data: List[Dict],
    filename: str,
) -> bool:
    """
    Write monthly data to DynamoDB for historical tracking.
    
    Args:
        branch_code: 3-letter code (e.g., "IMG")
        branch_name: Full name (e.g., "Imaginon")
        year_month_data: List of dicts with year_month, attendance, programs
        filename: Source filename for audit trail
    
    Returns:
        True if successful, False otherwise
    """
    table_name = _env("DYNAMODB_TABLE", "programming-data")
    
    try:
        table = dynamodb.Table(table_name)
        
        with table.batch_writer(batch_size=25) as batch:
            for data in year_month_data:
                item = {
                    "branch_code": branch_code,
                    "year_month": data["year_month"],  # YYYY-MM format
                    "year_month_gsi": data["year_month"],  # For GSI
                    "branch_name": branch_name,
                    "attendance": data["attendance"],
                    "programs": data["programs"],
                    "virtual_attendance": data.get("virtual_attendance", 0),
                    "created_date": datetime.now(timezone.utc).isoformat(),
                    "data_source_file": filename,
                }
                
                logger.info(
                    "Writing to DynamoDB: %s %s (attendance=%d, programs=%d)",
                    branch_code,
                    data["year_month"],
                    data["attendance"],
                    data["programs"],
                )
                
                batch.put_item(Item=item)

        logger.info("Successfully wrote %d records to DynamoDB for %s", len(year_month_data), branch_code)
        return True
        
    except Exception as exc:
        logger.exception("Error writing to DynamoDB: %s", str(exc))
        return False


def write_pdf_data_to_dynamodb(
    branch_code: str,
    branch_name: str,
    pdf_data: Dict,
    filename: str,
) -> bool:
    """
    Upsert a single month's PDF report data into DynamoDB.

    Uses update_item so that uploading an In-House report does not overwrite
    existing Outreach data for the same branch/month, and vice versa.

    In-House fields : attendance, programs
    Outreach fields : outreach_attendance, outreach_programs
    """
    table_name = _env("DYNAMODB_TABLE", "programming-data")

    try:
        table = dynamodb.Table(table_name)
        year_month = pdf_data["year_month"]
        now = datetime.now(timezone.utc).isoformat()

        if pdf_data["is_outreach"]:
            update_expr = (
                "SET branch_name = :name, year_month_gsi = :ym, "
                "outreach_programs = :p, outreach_attendance = :a, "
                "last_updated = :now, data_source_outreach = :file, "
                "created_date = if_not_exists(created_date, :now)"
            )
            attr_values = {
                ":name": branch_name,
                ":ym": year_month,
                ":p": pdf_data["programs"],
                ":a": pdf_data["attendance"],
                ":now": now,
                ":file": filename,
            }
        else:
            update_expr = (
                "SET branch_name = :name, year_month_gsi = :ym, "
                "attendance = :a, programs = :p, "
                "last_updated = :now, data_source_file = :file, "
                "created_date = if_not_exists(created_date, :now)"
            )
            attr_values = {
                ":name": branch_name,
                ":ym": year_month,
                ":a": pdf_data["attendance"],
                ":p": pdf_data["programs"],
                ":now": now,
                ":file": filename,
            }

        table.update_item(
            Key={"branch_code": branch_code, "year_month": year_month},
            UpdateExpression=update_expr,
            ExpressionAttributeValues=attr_values,
        )

        logger.info(
            "DynamoDB upsert: %s %s (%s) programs=%d attendance=%d",
            branch_code, year_month,
            "outreach" if pdf_data["is_outreach"] else "in-house",
            pdf_data["programs"], pdf_data["attendance"],
        )
        return True

    except Exception as exc:
        logger.exception("Error writing PDF data to DynamoDB: %s", str(exc))
        return False


def write_session_data_to_dynamodb(
    branch_code: str,
    branch_name: str,
    sessions: List[Dict],
    filename: str,
    year_month: str,
) -> int:
    """
    Write per-session rows extracted from a PDF report to the program_sessions table.

    Each item's sort key is  date#program_name#facilitator[#outreach_site]
    so that re-uploading the same PDF is idempotent (put_item overwrites).

    Returns the number of session records written (0 on failure or empty input).
    """
    table_name = _env("PROGRAM_SESSIONS_TABLE", "program-sessions")
    if not sessions:
        return 0

    try:
        table = dynamodb.Table(table_name)
        now = datetime.now(timezone.utc).isoformat()

        with table.batch_writer(batch_size=25) as batch:
            for s in sessions:
                outreach_site = s.get("outreach_site") or ""
                report_type = "outreach" if outreach_site else "in-house"

                key_parts = [s["program_date"], s["program_name"], s["primary_facilitator"]]
                if outreach_site:
                    key_parts.append(outreach_site)
                session_key = "#".join(key_parts)

                item: Dict[str, Any] = {
                    "branch_code": branch_code,
                    "session_key": session_key,
                    "program_date": s["program_date"],
                    "primary_facilitator": s["primary_facilitator"],
                    "program_name": s["program_name"],
                    "num_programs": s["num_programs"],
                    "total_attendance": s["total_attendance"],
                    "report_type": report_type,
                    "year_month": year_month,
                    "branch_name": branch_name,
                    "data_source_file": filename,
                    "created_at": now,
                }
                if outreach_site:
                    item["outreach_site"] = outreach_site

                batch.put_item(Item=item)

        logger.info(
            "Wrote %d session records to DynamoDB for %s %s",
            len(sessions), branch_code, year_month,
        )
        return len(sessions)

    except Exception as exc:
        logger.exception("Error writing session data to DynamoDB: %s", str(exc))
        return 0


# ── Handlers ─────────────────────────────────────────────────────────────────

def handle_s3_event(event: dict) -> dict:
    """S3 PUT trigger. Downloads .xlsx, parses it, writes to S3 and DynamoDB."""
    try:
        record = event["Records"][0]["s3"]
        source_bucket = record["bucket"]["name"]
        source_key = unquote_plus(record["object"]["key"])

        logger.info("Processing s3://%s/%s", source_bucket, source_key)

        filename = source_key.split("/")[-1]
        is_pdf = filename.lower().endswith(".pdf")

        # Branch code: PDFs use full name in filename; xlsx use 3-letter prefix
        branch_code = (
            _extract_branch_code_from_name(filename)
            if is_pdf
            else _extract_branch_code(filename)
        )

        if not branch_code:
            logger.error("Cannot extract branch code from filename: %s", filename)
            return {
                "statusCode": 400,
                "body": json.dumps({"error": "Cannot determine branch from filename"}),
            }

        branch_name = _get_branch_name(branch_code)
        file_bytes = read_s3(source_bucket, source_key)

        # ── PDF path ──────────────────────────────────────────────────────────
        if is_pdf:
            pdf_data = parse_pdf_report(file_bytes)
            if not pdf_data:
                return {
                    "statusCode": 400,
                    "body": json.dumps({"error": "Could not parse PDF report"}),
                }

            dynamodb_success = write_pdf_data_to_dynamodb(
                branch_code, branch_name, pdf_data, filename
            )
            sessions_written = write_session_data_to_dynamodb(
                branch_code, branch_name,
                pdf_data.get("sessions", []),
                filename, pdf_data["year_month"],
            )

            logger.info("Processed PDF %s for branch %s", filename, branch_code)
            return {
                "statusCode": 200,
                "body": json.dumps({
                    "message": "PDF programming report processed",
                    "branch": branch_code,
                    "yearMonth": pdf_data["year_month"],
                    "reportType": "outreach" if pdf_data["is_outreach"] else "in-house",
                    "programs": pdf_data["programs"],
                    "attendance": pdf_data["attendance"],
                    "sessionsWritten": sessions_written,
                    "dynamodbSuccess": dynamodb_success,
                }),
            }

        # ── Excel path ────────────────────────────────────────────────────────
        parsed = parse_workbook(file_bytes)

        if not parsed["data"]:
            logger.warning("No valid data found in file")
            return {
                "statusCode": 400,
                "body": json.dumps({"error": "No valid data in workbook"}),
            }

        # Write to DynamoDB (historical data)
        dynamodb_success = write_programming_data_to_dynamodb(
            branch_code,
            branch_name,
            parsed["year_month_data"],
            filename,
        )

        if not dynamodb_success:
            logger.warning("DynamoDB write failed, but continuing to S3")

        # Write to S3: processed/programming/{BRANCH_CODE}.json
        output_payload = {
            "branch": branch_code,
            "branchName": branch_name,
            "data": parsed["data"],
            "months": parsed["months"],
            "attendance": parsed["attendance"],
            "programs": parsed["programs"],
            "lastUpdated": datetime.now(timezone.utc).isoformat(),
            "dataFound": True,
        }

        dest_bucket = _env("PROCESSED_BUCKET", source_bucket)
        dest_key = f"{PROCESSED_DIR}/{branch_code}.json"
        write_json_s3(dest_bucket, dest_key, output_payload)

        logger.info("Successfully processed %s for branch %s", filename, branch_code)
        return {
            "statusCode": 200,
            "body": json.dumps({
                "message": "Programming data processed",
                "branch": branch_code,
                "dataPoints": len(parsed["data"]),
                "dynamodbSuccess": dynamodb_success,
            }),
        }

    except Exception as exc:
        logger.exception("Error processing S3 event")
        return {
            "statusCode": 500,
            "body": json.dumps({"error": str(exc)}),
        }


# ── Lambda Handler Entrypoint ────────────────────────────────────────────────

def lambda_handler(event, context):
    """S3 PUT trigger only — parses uploaded programming files and writes to DynamoDB."""
    logger.info("Event: %s", json.dumps(event, default=str)[:500])

    if "Records" not in event:
        logger.error("Non-S3 event received — this lambda is an S3 trigger only")
        return {
            "statusCode": 400,
            "body": json.dumps({"error": "This lambda only handles S3 events"}),
        }

    return handle_s3_event(event)
