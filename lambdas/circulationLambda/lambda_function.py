"""
Library Analytics – Circulation Statistics Lambda (with Breakdown Support)

Parses FY circulation .xlsm files uploaded to S3 and serves processed
graph data to the frontend dashboard via API Gateway.

Features:
  - Extracts per-category breakdowns (e.g., Juvenile Fiction + Juvenile Non-Fiction)
  - Returns data with breakdown field for tooltip visualization
  - Supports branch-level filtering
  - Category renamed: "Juvenile Fiction" → "Juvenile"

Trigger:  S3 PUT on uploads/circulation/*.xlsm
API:      GET /circulation?branch={name}

Response includes:
  - branches: list of all branches with data
  - data: CirculationDataPoint[] with breakdown field
"""

import json
import logging
import os
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any
from urllib.parse import unquote_plus

import boto3
import openpyxl
from botocore.exceptions import ClientError

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client("s3")

# ── Constants ────────────────────────────────────────────────────────────────

FY_MONTHS = [
    "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER",
    "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
]

MONTH_ABBR = {
    "JULY": "Jul", "AUGUST": "Aug", "SEPTEMBER": "Sep",
    "OCTOBER": "Oct", "NOVEMBER": "Nov", "DECEMBER": "Dec",
    "JANUARY": "Jan", "FEBRUARY": "Feb", "MARCH": "Mar",
    "APRIL": "Apr", "MAY": "May", "JUNE": "Jun",
}

# ── COLUMN INDEX MAPPING (0-based) ───────────────────────────────────────
#
# The spreadsheet columns are structured as:
#   Col 0: Department name
#   Col 1: blank
#   Cols 2-4: Adult Fiction, Adult Non-Fiction, TOTAL ADULT
#   Col 5: blank
#   Cols 6-8: Juvenile Fiction, Juvenile Non-Fiction, TOTAL JUVENILE
#   Col 9: blank
#   Cols 10-12: YA Fiction, YA Non-Fiction, TOTAL YA
#   Col 13: blank
#   Cols 14,15,16: Book Downloads, blank, TOTAL BOOKS
#   Col 17: blank
#   Cols 18-21: Audio, Visual, blank, TOTAL NONPRINT
#   Col 22: blank
#   Col 23: TOTAL PRINT & NONPRINT (Grand Total)

COL_ADULT_FICTION      = 2
COL_ADULT_NONFICTION   = 3
COL_TOTAL_ADULT        = 4

COL_JUV_FICTION        = 6
COL_JUV_NONFICTION     = 7
COL_TOTAL_JUVENILE     = 8

COL_YA_FICTION         = 10
COL_YA_NONFICTION      = 11
COL_TOTAL_YA           = 12

COL_BOOK_DOWNLOADS     = 14
COL_TOTAL_BOOKS        = 16

COL_AUDIO              = 18
COL_VISUAL             = 19
COL_TOTAL_NONPRINT     = 21

COL_GRAND_TOTAL        = 23

# Category definitions with breakdown support
CATEGORY_DEFINITIONS = [
    {
        "category": "Juvenile",
        "total_col": COL_TOTAL_JUVENILE,
        "breakdown": {
            "Juvenile Fiction": COL_JUV_FICTION,
            "Juvenile Non-Fiction": COL_JUV_NONFICTION,
        },
    },
    {
        "category": "Young Adult",
        "total_col": COL_TOTAL_YA,
        "breakdown": {
            "Young Adult Fiction": COL_YA_FICTION,
            "Young Adult Non-Fiction": COL_YA_NONFICTION,
        },
    },
    {
        "category": "Adult",
        "total_col": COL_TOTAL_ADULT,
        "breakdown": {
            "Adult Fiction": COL_ADULT_FICTION,
            "Adult Non-Fiction": COL_ADULT_NONFICTION,
        },
    },
    {
        "category": "Non-Print",
        "total_col": COL_TOTAL_NONPRINT,
        "breakdown": {
            "Audio": COL_AUDIO,
            "Visual": COL_VISUAL,
        },
    },
    {
        "category": "Total Circulation",
        "total_col": COL_GRAND_TOTAL,
        "breakdown": {
            "Total Books": COL_TOTAL_BOOKS,
            "Total NonPrint": COL_TOTAL_NONPRINT,
        },
    },
]

PROCESSED_KEY = "processed/circulation_data.json"

# Branch names to skip (system totals, headers, etc.)
SKIP_DEPARTMENTS = {
    "Department", "Charlotte Mecklenburg Library", "CIRCULATION", "Page 1",
    "Total", "TOTAL ADULT", "TOTAL JUVENILE", "TOTAL YOUNG ADULT", 
    "TOTAL NONPRINT", "TOTAL BOOKS", "TOTAL PRINT & NONPRINT",
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def _env(name: str, fallback: str = "") -> str:
    return os.environ.get(name, fallback)


def _safe_int(val: Any) -> int:
    if val is None:
        return 0
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _is_skip_department(name: str) -> bool:
    """Check if a department name should be skipped (system total, header, etc.)"""
    if not name or not isinstance(name, str):
        return True
    name_upper = name.upper().strip()
    return any(skip.upper() in name_upper for skip in SKIP_DEPARTMENTS)


# ── Excel Parsing ────────────────────────────────────────────────────────────

def _find_header_row(sheet) -> int:
    """Find the row containing 'Department' header."""
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if row and row[0] and str(row[0]).strip() == "Department":
            return i
    return -1


def _extract_departments(sheet) -> list[str]:
    """Extract all unique branch/department names from a sheet."""
    header_idx = _find_header_row(sheet)
    if header_idx < 0:
        return []

    departments = set()
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        if row and row[0]:
            dept_name = str(row[0]).strip()
            if not _is_skip_department(dept_name):
                departments.add(dept_name)

    return sorted(list(departments))


def _find_totals_row(sheet) -> tuple | None:
    """Find the system-wide TOTAL row (for backward compatibility)."""
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] and str(row[0]).strip() == "Total":
            return row
    return None


def _extract_branch_data_with_breakdown(sheet, branch_name: str) -> dict | None:
    """
    Extract data for a specific branch from a sheet.
    Returns dict with all category data including breakdowns.
    """
    header_idx = _find_header_row(sheet)
    if header_idx < 0:
        return None

    for row in sheet.iter_rows(values_only=True):
        if row and row[0] and str(row[0]).strip() == branch_name:
            data = {}
            for cat_def in CATEGORY_DEFINITIONS:
                total_val = _safe_int(row[cat_def["total_col"]])
                data[cat_def["category"]] = {
                    "total": total_val,
                    "breakdown": {
                        label: _safe_int(row[col])
                        for label, col in cat_def["breakdown"].items()
                    }
                }
            return data
    return None


def parse_workbook(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Parse the .xlsm and return:
      1. List of month records in fiscal-year order with branch data (with breakdowns)
      2. List of unique branch names across all months

    Each month record: {
      month_name, display_month, year,
      system_totals: {category: {total: int, breakdown: {label: int}}},
      branches: {branch_name: {category: {total: int, breakdown: {label: int}}}}
    }
    """
    wb = openpyxl.load_workbook(
        BytesIO(file_bytes), read_only=True, data_only=True, keep_links=False,
    )

    result = []
    all_branches = set()

    for month_name in FY_MONTHS:
        if month_name not in wb.sheetnames:
            continue

        sheet = wb[month_name]
        total_row = _find_totals_row(sheet)
        if total_row is None:
            continue

        grand_total = _safe_int(total_row[COL_GRAND_TOTAL])
        if grand_total == 0:
            continue

        # Extract system totals with breakdowns
        year = _detect_year(sheet, month_name)
        abbr = MONTH_ABBR[month_name]
        display = f"{abbr} {year}"

        system_totals = {}
        for cat_def in CATEGORY_DEFINITIONS:
            total_val = _safe_int(total_row[cat_def["total_col"]])
            system_totals[cat_def["category"]] = {
                "total": total_val,
                "breakdown": {
                    label: _safe_int(total_row[col])
                    for label, col in cat_def["breakdown"].items()
                }
            }

        # Extract branch data with breakdowns
        branches = {}
        for dept_name in _extract_departments(sheet):
            dept_data = _extract_branch_data_with_breakdown(sheet, dept_name)
            if dept_data:
                grand = dept_data.get("Total Circulation", {}).get("total", 0)
                if grand > 0:
                    branches[dept_name] = dept_data
                    all_branches.add(dept_name)

        result.append({
            "month_name": month_name,
            "display_month": display,
            "year": year,
            "system_totals": system_totals,
            "branches": branches,
        })

    wb.close()
    return result[-12:], sorted(list(all_branches))


def _detect_year(sheet, month_name: str) -> int:
    """Extract the calendar year from the sheet header."""
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and month_name in cell.upper():
                for token in cell.split():
                    if token.isdigit() and len(token) == 4:
                        return int(token)
    return datetime.now(timezone.utc).year


# ── Payload Builder ──────────────────────────────────────────────────────────

def build_circulation_data(months: list[dict], branch_filter: str = "") -> dict:
    """
    Convert parsed month records into CirculationDataPoint[] with breakdowns.
    
    If branch_filter is empty or "System", use system totals.
    If branch_filter is a branch name, use that branch's data.
    """
    now = datetime.now(timezone.utc).isoformat()
    points: list[dict] = []
    all_branches = set()

    # Collect all unique branches
    for month_rec in months:
        all_branches.update(month_rec["branches"].keys())

    # Extract data based on filter
    for month_rec in months:
        if not branch_filter or branch_filter == "System":
            # Use system totals
            for category, cat_data in month_rec["system_totals"].items():
                point = {
                    "category": category,
                    "month": month_rec["display_month"],
                    "year": month_rec["year"],
                    "circulation": cat_data["total"],
                }
                if cat_data.get("breakdown"):
                    point["breakdown"] = cat_data["breakdown"]
                points.append(point)
        elif branch_filter in month_rec["branches"]:
            # Use branch data
            for category, cat_data in month_rec["branches"][branch_filter].items():
                point = {
                    "category": category,
                    "month": month_rec["display_month"],
                    "year": month_rec["year"],
                    "circulation": cat_data["total"],
                }
                if cat_data.get("breakdown"):
                    point["breakdown"] = cat_data["breakdown"]
                points.append(point)

    return {
        "data": points,
        "branches": sorted(list(all_branches)),
        "selectedBranch": branch_filter or "System",
        "lastUpdated": now,
        "totalRecords": len(points),
    }


# ── S3 I/O ───────────────────────────────────────────────────────────────────

def read_s3(bucket: str, key: str) -> bytes:
    return s3.get_object(Bucket=bucket, Key=key)["Body"].read()


def write_json_s3(bucket: str, key: str, payload: dict) -> None:
    s3.put_object(
        Bucket=bucket,
        Key=key,
        Body=json.dumps(payload, indent=2),
        ContentType="application/json",
    )
    logger.info("Wrote s3://%s/%s", bucket, key)


def read_json_s3(bucket: str, key: str) -> dict:
    return json.loads(s3.get_object(Bucket=bucket, Key=key)["Body"].read())


# ── API response helpers ─────────────────────────────────────────────────────

def _api_ok(data: Any) -> dict:
    return _api_response(200, {
        "success": True,
        "data": data,
        "error": None,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "requestId": f"req_{uuid.uuid4().hex[:12]}",
    })


def _api_err(status: int, code: str, message: str) -> dict:
    return _api_response(status, {
        "success": False,
        "data": None,
        "error": {"code": code, "message": message},
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "requestId": f"req_{uuid.uuid4().hex[:12]}",
    })


def _api_response(status: int, body: dict) -> dict:
    return {
        "statusCode": status,
        "headers": {
            "Content-Type": "application/json",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "Content-Type, Authorization",
        },
        "body": json.dumps(body),
    }


# ── Handlers ─────────────────────────────────────────────────────────────────

def handle_s3_event(event: dict) -> dict:
    """S3 PUT trigger. Downloads .xlsm, parses it, writes processed JSON."""
    record = event["Records"][0]["s3"]
    source_bucket = record["bucket"]["name"]
    source_key = unquote_plus(record["object"]["key"])

    logger.info("Processing s3://%s/%s", source_bucket, source_key)

    file_bytes = read_s3(source_bucket, source_key)
    months, branches = parse_workbook(file_bytes)

    if not months:
        logger.error("No valid month data found")
        return {"statusCode": 400, "body": "No circulation data in workbook"}

    # Store raw months + branch data so the API can filter by any branch at request time
    stored = {
        "months": months,
        "branches": branches,
        "lastUpdated": datetime.now(timezone.utc).isoformat(),
    }

    dest_bucket = _env("PROCESSED_BUCKET", source_bucket)
    dest_key = _env("PROCESSED_KEY", PROCESSED_KEY)
    write_json_s3(dest_bucket, dest_key, stored)

    return {
        "statusCode": 200,
        "body": json.dumps({
            "message": "Circulation data processed with breakdowns",
            "monthsFound": len(months),
            "branchesFound": len(branches),
        }),
    }


def handle_api_request(event: dict) -> dict:
    """GET /circulation?branch={name}"""
    bucket = _env("PROCESSED_BUCKET")
    if not bucket:
        return _api_err(500, "CONFIG_ERROR", "PROCESSED_BUCKET not set")

    key = _env("PROCESSED_KEY", PROCESSED_KEY)

    try:
        stored = read_json_s3(bucket, key)
    except ClientError as exc:
        code = exc.response["Error"]["Code"]
        if code in ("NoSuchKey", "404", "AccessDenied", "403"):
            return _api_err(404, "NOT_FOUND", "No data available. Upload a file first.")
        logger.exception("S3 error reading processed data")
        return _api_err(500, "INTERNAL_ERROR", f"Storage error: {code}")
    except Exception as exc:
        logger.exception("Error reading processed data")
        return _api_err(500, "INTERNAL_ERROR", str(exc))

    branch_filter = (event.get("queryStringParameters") or {}).get("branch", "System")
    payload = build_circulation_data(stored["months"], branch_filter)
    return _api_ok(payload)


def lambda_handler(event, context):
    """Main Lambda entry point. Routes S3 events vs API Gateway requests."""
    logger.info("Event: %s", json.dumps(event))

    # Determine event type
    if "Records" in event:
        return handle_s3_event(event)
    else:
        return handle_api_request(event)
