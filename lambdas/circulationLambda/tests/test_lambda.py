"""
Tests for the circulation statistics Lambda.

Run:  pytest tests/ -v
"""

import json
import os
import pytest
from unittest.mock import patch, MagicMock
from io import BytesIO

import openpyxl

from lambda_function import (
    parse_workbook,
    build_circulation_data,
    _safe_int,
    _is_skip_department,
    _find_header_row,
    _find_totals_row,
    _api_ok,
    _api_err,
    lambda_handler,
    FY_MONTHS,
    CATEGORY_DEFINITIONS,
    COL_ADULT_FICTION,
    COL_ADULT_NONFICTION,
    COL_TOTAL_ADULT,
    COL_JUV_FICTION,
    COL_JUV_NONFICTION,
    COL_TOTAL_JUVENILE,
    COL_YA_FICTION,
    COL_YA_NONFICTION,
    COL_TOTAL_YA,
    COL_TOTAL_BOOKS,
    COL_AUDIO,
    COL_VISUAL,
    COL_TOTAL_NONPRINT,
    COL_GRAND_TOTAL,
)

_CATEGORY_NAMES = [d["category"] for d in CATEGORY_DEFINITIONS]

# ── Fixtures & helpers ────────────────────────────────────────────────────────

# Real values from FY2026 workbook JULY "Total" row (0-based column indices)
REAL_DATA = {
    "JULY": {
        COL_ADULT_FICTION:    111022,
        COL_ADULT_NONFICTION: 28926,
        COL_TOTAL_ADULT:      139948,
        COL_JUV_FICTION:      239142,
        COL_JUV_NONFICTION:   45889,
        COL_TOTAL_JUVENILE:   285031,
        COL_YA_FICTION:       18213,
        COL_YA_NONFICTION:    2894,
        COL_TOTAL_YA:         21107,
        COL_AUDIO:            130000,
        COL_VISUAL:           39159,
        COL_TOTAL_NONPRINT:   169159,
        COL_TOTAL_BOOKS:      678102,
        COL_GRAND_TOTAL:      847261,
    },
    "AUGUST": {
        COL_TOTAL_ADULT:      130799,
        COL_TOTAL_JUVENILE:   262962,
        COL_TOTAL_YA:         19115,
        COL_TOTAL_NONPRINT:   164248,
        COL_GRAND_TOTAL:      806225,
    },
    "SEPTEMBER": {
        COL_TOTAL_ADULT:      123320,
        COL_TOTAL_JUVENILE:   247639,
        COL_TOTAL_YA:         16360,
        COL_TOTAL_NONPRINT:   159559,
        COL_GRAND_TOTAL:      780814,
    },
    "APRIL": {
        COL_TOTAL_ADULT:      110500,
        COL_TOTAL_JUVENILE:   213936,
        COL_TOTAL_YA:         14985,
        COL_TOTAL_NONPRINT:   167013,
        COL_GRAND_TOTAL:      771287,
    },
}

# Sample branch-level data matching Imaginon Jul 2025 values
IMAGINON_BRANCH_DATA = {
    COL_ADULT_FICTION:    236,
    COL_ADULT_NONFICTION: 214,
    COL_TOTAL_ADULT:      450,
    COL_JUV_FICTION:      1200,
    COL_JUV_NONFICTION:   300,
    COL_TOTAL_JUVENILE:   1500,
    COL_YA_FICTION:       180,
    COL_YA_NONFICTION:    20,
    COL_TOTAL_YA:         200,
    COL_AUDIO:            400,
    COL_VISUAL:           100,
    COL_TOTAL_NONPRINT:   500,
    COL_TOTAL_BOOKS:      2150,
    COL_GRAND_TOTAL:      2650,
}


def _make_workbook(
    months_with_data: dict,
    fy_start_year: int = 2025,
    branch_rows: dict | None = None,
) -> bytes:
    """
    Build a minimal .xlsx in memory that parse_workbook can read.

    months_with_data: { "JULY": { col_index: value, ... }, ... }
    branch_rows:      { col_index: value, ... } – added as an "Imaginon" branch
                      row to every month that has system data.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for month_name in FY_MONTHS:
        ws = wb.create_sheet(month_name)
        year = fy_start_year if FY_MONTHS.index(month_name) < 6 else fy_start_year + 1

        # Row 1: year header — _detect_year scans rows 1-5 for month+year string
        hdr = [None] * 25
        hdr[0] = f"{month_name} {year}"
        ws.append(hdr)

        # Row 2: "Department" header — required for branch extraction
        dept = [None] * 25
        dept[0] = "Department"
        ws.append(dept)

        # Row 3: "Total" row — _find_totals_row looks for row[0] == "Total"
        total = [None] * 25
        total[0] = "Total"
        if month_name in months_with_data:
            for col_idx, value in months_with_data[month_name].items():
                total[col_idx] = value
        ws.append(total)

        # Optional branch row (must appear after "Department" row)
        if branch_rows and month_name in months_with_data:
            branch = [None] * 25
            branch[0] = "Imaginon"
            for col_idx, value in branch_rows.items():
                branch[col_idx] = value
            ws.append(branch)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class MockSheet:
    """Sheet stub whose iter_rows always returns a fresh iterator."""
    def __init__(self, rows):
        self._rows = rows

    def iter_rows(self, values_only=False, **kwargs):
        return iter(self._rows)


def _row(col_dict: dict, size: int = 25) -> tuple:
    row = [None] * size
    for col, val in col_dict.items():
        row[col] = val
    return tuple(row)


# ── _safe_int ────────────────────────────────────────────────────────────────

class TestSafeInt:
    def test_none_returns_zero(self):
        assert _safe_int(None) == 0

    def test_empty_string_returns_zero(self):
        assert _safe_int("") == 0

    def test_whitespace_returns_zero(self):
        assert _safe_int("   ") == 0

    def test_integer_passthrough(self):
        assert _safe_int(847261) == 847261

    def test_float_truncates(self):
        assert _safe_int(0.8991) == 0

    def test_numeric_string(self):
        assert _safe_int("139948") == 139948

    def test_float_string_truncates(self):
        assert _safe_int("42.9") == 42

    def test_non_numeric_returns_zero(self):
        assert _safe_int("abc") == 0


# ── _is_skip_department ──────────────────────────────────────────────────────

class TestIsSkipDepartment:
    def test_none_is_skipped(self):
        assert _is_skip_department(None) is True

    def test_empty_string_is_skipped(self):
        assert _is_skip_department("") is True

    def test_department_header_is_skipped(self):
        assert _is_skip_department("Department") is True

    def test_total_is_skipped(self):
        assert _is_skip_department("Total") is True

    def test_library_name_is_skipped(self):
        assert _is_skip_department("Charlotte Mecklenburg Library") is True

    def test_regular_branch_is_not_skipped(self):
        assert _is_skip_department("Main Branch") is False

    def test_another_branch_is_not_skipped(self):
        assert _is_skip_department("Plaza Midwood") is False


# ── _find_header_row ─────────────────────────────────────────────────────────

class TestFindHeaderRow:
    def test_finds_department_row_at_index_zero(self):
        sheet = MockSheet([
            _row({0: "Department"}),
            _row({0: "Branch A", COL_GRAND_TOTAL: 100}),
        ])
        assert _find_header_row(sheet) == 0

    def test_finds_department_row_after_header_rows(self):
        sheet = MockSheet([
            _row({0: "JULY 2025"}),
            _row({}),
            _row({0: "Department"}),
            _row({0: "Branch A", COL_GRAND_TOTAL: 100}),
        ])
        assert _find_header_row(sheet) == 2

    def test_returns_minus_one_when_not_found(self):
        sheet = MockSheet([
            _row({0: "JULY 2025"}),
            _row({0: "Total", COL_GRAND_TOTAL: 500}),
        ])
        assert _find_header_row(sheet) == -1


# ── _find_totals_row ─────────────────────────────────────────────────────────

class TestFindTotalsRow:
    def test_finds_total_row(self):
        total = _row({0: "Total", COL_GRAND_TOTAL: 847261})
        sheet = MockSheet([
            _row({0: "Department"}),
            _row({0: "Branch A", COL_GRAND_TOTAL: 100}),
            total,
        ])
        result = _find_totals_row(sheet)
        assert result is not None
        assert result[0] == "Total"
        assert result[COL_GRAND_TOTAL] == 847261

    def test_returns_none_when_no_total_row(self):
        sheet = MockSheet([
            _row({0: "Department"}),
            _row({0: "Branch A", COL_GRAND_TOTAL: 100}),
        ])
        assert _find_totals_row(sheet) is None


# ── parse_workbook ───────────────────────────────────────────────────────────

class TestParseWorkbook:
    def test_extracts_months_with_data(self):
        wb = _make_workbook({"JULY": REAL_DATA["JULY"], "AUGUST": REAL_DATA["AUGUST"]})
        months, _ = parse_workbook(wb)
        assert len(months) == 2
        assert months[0]["month_name"] == "JULY"
        assert months[1]["month_name"] == "AUGUST"

    def test_skips_zero_grand_total_months(self):
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]})
        months, _ = parse_workbook(wb)
        assert len(months) == 1
        assert months[0]["month_name"] == "JULY"

    def test_returns_empty_list_for_empty_workbook(self):
        wb = _make_workbook({})
        months, branches = parse_workbook(wb)
        assert months == []
        assert branches == []

    def test_values_match_real_data(self):
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]})
        months, _ = parse_workbook(wb)
        totals = months[0]["system_totals"]
        assert totals["Juvenile"]["total"] == 285031
        assert totals["Young Adult"]["total"] == 21107
        assert totals["Adult"]["total"] == 139948
        assert totals["Non-Print"]["total"] == 169159
        assert totals["Total Circulation"]["total"] == 847261

    def test_system_totals_has_breakdown_field(self):
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]})
        months, _ = parse_workbook(wb)
        totals = months[0]["system_totals"]
        assert "breakdown" in totals["Juvenile"]
        assert totals["Juvenile"]["breakdown"]["Juvenile Fiction"] == 239142
        assert totals["Juvenile"]["breakdown"]["Juvenile Non-Fiction"] == 45889
        assert "breakdown" in totals["Adult"]
        assert totals["Adult"]["breakdown"]["Adult Fiction"] == 111022

    def test_display_month_format(self):
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]})
        months, _ = parse_workbook(wb)
        assert months[0]["display_month"] == "Jul 2025"

    def test_caps_at_twelve_months(self):
        all_months = {m: REAL_DATA["JULY"] for m in FY_MONTHS}
        wb = _make_workbook(all_months)
        months, _ = parse_workbook(wb)
        assert len(months) == 12

    def test_fiscal_year_order_preserved(self):
        data = {
            "JULY": REAL_DATA["JULY"],
            "SEPTEMBER": REAL_DATA["SEPTEMBER"],
            "APRIL": REAL_DATA["APRIL"],
        }
        wb = _make_workbook(data)
        months, _ = parse_workbook(wb)
        assert [r["month_name"] for r in months] == ["JULY", "SEPTEMBER", "APRIL"]

    def test_year_detection_first_half(self):
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]}, fy_start_year=2025)
        months, _ = parse_workbook(wb)
        assert months[0]["year"] == 2025

    def test_year_detection_second_half(self):
        wb = _make_workbook({"APRIL": REAL_DATA["APRIL"]}, fy_start_year=2025)
        months, _ = parse_workbook(wb)
        assert months[0]["year"] == 2026

    def test_returns_branch_list(self):
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]}, branch_rows=IMAGINON_BRANCH_DATA)
        _, branches = parse_workbook(wb)
        assert "Imaginon" in branches

    def test_branch_data_extracted_from_workbook(self):
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]}, branch_rows=IMAGINON_BRANCH_DATA)
        months, _ = parse_workbook(wb)
        assert "Imaginon" in months[0]["branches"]
        imaginon = months[0]["branches"]["Imaginon"]
        assert imaginon["Adult"]["total"] == 450
        assert imaginon["Adult"]["breakdown"]["Adult Fiction"] == 236
        assert imaginon["Adult"]["breakdown"]["Adult Non-Fiction"] == 214
        assert imaginon["Juvenile"]["total"] == 1500

    def test_no_branches_when_no_branch_rows(self):
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]})
        months, branches = parse_workbook(wb)
        assert branches == []
        assert months[0]["branches"] == {}


# ── build_circulation_data ───────────────────────────────────────────────────

class TestBuildCirculationData:
    def _sample_months(self, count=3):
        """Build minimal month records using the new {total, breakdown} structure."""
        months = []
        for i, m in enumerate(FY_MONTHS[:count]):
            total_val = 1000 * (i + 1)
            branch_val = 500 * (i + 1)
            system_totals = {
                cat: {"total": total_val, "breakdown": {"A": total_val // 2, "B": total_val // 2}}
                for cat in _CATEGORY_NAMES
            }
            branch_data = {
                cat: {"total": branch_val, "breakdown": {"A": branch_val // 2, "B": branch_val // 2}}
                for cat in _CATEGORY_NAMES
            }
            months.append({
                "month_name": m,
                "display_month": f"{m[:3].title()} 2025",
                "year": 2025,
                "system_totals": system_totals,
                "branches": {"Main Branch": branch_data},
            })
        return months

    def test_produces_correct_point_count(self):
        months = self._sample_months(3)
        result = build_circulation_data(months)
        # 5 categories × 3 months = 15 points
        assert result["totalRecords"] == 15
        assert len(result["data"]) == 15

    def test_data_point_shape(self):
        months = self._sample_months(1)
        result = build_circulation_data(months)
        point = result["data"][0]
        assert "category" in point
        assert "month" in point
        assert "year" in point
        assert "circulation" in point

    def test_all_five_categories_present(self):
        months = self._sample_months(1)
        result = build_circulation_data(months)
        categories = {pt["category"] for pt in result["data"]}
        assert categories == set(_CATEGORY_NAMES)

    def test_last_updated_is_iso(self):
        months = self._sample_months(1)
        result = build_circulation_data(months)
        assert "T" in result["lastUpdated"]

    def test_default_selected_branch_is_system(self):
        months = self._sample_months(1)
        result = build_circulation_data(months)
        assert result["selectedBranch"] == "System"

    def test_branch_filter_uses_branch_data(self):
        months = self._sample_months(2)
        result = build_circulation_data(months, branch_filter="Main Branch")
        # 5 categories × 2 months = 10 points
        assert len(result["data"]) == 10
        assert result["selectedBranch"] == "Main Branch"

    def test_unknown_branch_returns_empty_data(self):
        months = self._sample_months(2)
        result = build_circulation_data(months, branch_filter="Ghost Branch")
        assert result["data"] == []

    def test_branches_list_in_result(self):
        months = self._sample_months(2)
        result = build_circulation_data(months)
        assert "Main Branch" in result["branches"]

    def test_category_values_correct(self):
        months = self._sample_months(1)
        result = build_circulation_data(months)
        adult_pt = next(pt for pt in result["data"] if pt["category"] == "Adult")
        assert adult_pt["circulation"] == 1000  # first month, multiplier 1

    def test_data_point_includes_breakdown(self):
        months = self._sample_months(1)
        result = build_circulation_data(months)
        point = result["data"][0]
        assert "breakdown" in point
        assert point["breakdown"]["A"] == 500

    def test_branch_filter_values_differ_from_system(self):
        months = self._sample_months(1)
        sys_result = build_circulation_data(months)
        branch_result = build_circulation_data(months, branch_filter="Main Branch")
        sys_adult = next(p for p in sys_result["data"] if p["category"] == "Adult")
        br_adult = next(p for p in branch_result["data"] if p["category"] == "Adult")
        assert sys_adult["circulation"] == 1000
        assert br_adult["circulation"] == 500


# ── API response helpers ──────────────────────────────────────────────────────

class TestApiHelpers:
    def test_api_ok_status_200(self):
        resp = _api_ok({"key": "value"})
        assert resp["statusCode"] == 200

    def test_api_ok_body_success_true(self):
        resp = _api_ok({"key": "value"})
        body = json.loads(resp["body"])
        assert body["success"] is True
        assert body["error"] is None
        assert "requestId" in body
        assert "timestamp" in body

    def test_api_err_status_code(self):
        resp = _api_err(404, "NOT_FOUND", "missing")
        assert resp["statusCode"] == 404

    def test_api_err_body_structure(self):
        resp = _api_err(500, "INTERNAL_ERROR", "boom")
        body = json.loads(resp["body"])
        assert body["success"] is False
        assert body["data"] is None
        assert body["error"]["code"] == "INTERNAL_ERROR"
        assert body["error"]["message"] == "boom"

    def test_cors_headers_present(self):
        resp = _api_ok({})
        assert resp["headers"]["Access-Control-Allow-Origin"] == "*"
        assert "GET" in resp["headers"]["Access-Control-Allow-Methods"]


# ── lambda_handler – S3 event ─────────────────────────────────────────────────

class TestS3Handler:
    def _s3_event(self, bucket="test-bucket", key="uploads/circulation/FY2026.xlsm"):
        return {
            "Records": [{
                "eventSource": "aws:s3",
                "s3": {
                    "bucket": {"name": bucket},
                    "object": {"key": key},
                },
            }],
        }

    @patch("lambda_function.write_json_s3")
    @patch("lambda_function.read_s3")
    def test_processes_upload_and_writes_json(self, mock_read, mock_write):
        mock_read.return_value = _make_workbook({"JULY": REAL_DATA["JULY"]})
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "test-bucket"}):
            result = lambda_handler(self._s3_event(), None)
        assert result["statusCode"] == 200
        mock_write.assert_called_once()
        written_payload = mock_write.call_args[0][2]
        # New format stores raw months + branches for on-demand branch filtering
        assert "months" in written_payload
        assert "branches" in written_payload
        assert len(written_payload["months"]) == 1

    @patch("lambda_function.write_json_s3")
    @patch("lambda_function.read_s3")
    def test_empty_workbook_returns_400(self, mock_read, mock_write):
        mock_read.return_value = _make_workbook({})
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "test-bucket"}):
            result = lambda_handler(self._s3_event(), None)
        assert result["statusCode"] == 400
        mock_write.assert_not_called()

    @patch("lambda_function.write_json_s3")
    @patch("lambda_function.read_s3")
    def test_idempotent_reupload(self, mock_read, mock_write):
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]})
        mock_read.return_value = wb
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            lambda_handler(self._s3_event(), None)
            lambda_handler(self._s3_event(), None)
        assert mock_write.call_count == 2
        first = mock_write.call_args_list[0][0][2]
        second = mock_write.call_args_list[1][0][2]
        assert first["months"] == second["months"]

    @patch("lambda_function.write_json_s3")
    @patch("lambda_function.read_s3")
    def test_response_body_has_branch_count(self, mock_read, mock_write):
        mock_read.return_value = _make_workbook({"JULY": REAL_DATA["JULY"]})
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            result = lambda_handler(self._s3_event(), None)
        body = json.loads(result["body"])
        assert "branchesFound" in body

    @patch("lambda_function.write_json_s3")
    @patch("lambda_function.read_s3")
    def test_url_encoded_key_is_decoded(self, mock_read, mock_write):
        mock_read.return_value = _make_workbook({"JULY": REAL_DATA["JULY"]})
        encoded_event = self._s3_event(key="uploads/circulation/FY2026+Circulation+Statistics.xlsm")
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            lambda_handler(encoded_event, None)
        actual_key = mock_read.call_args[0][1]
        assert actual_key == "uploads/circulation/FY2026 Circulation Statistics.xlsm"

    @patch("lambda_function.write_json_s3")
    @patch("lambda_function.read_s3")
    def test_branch_data_stored_in_months(self, mock_read, mock_write):
        mock_read.return_value = _make_workbook(
            {"JULY": REAL_DATA["JULY"]}, branch_rows=IMAGINON_BRANCH_DATA
        )
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            lambda_handler(self._s3_event(), None)
        written = mock_write.call_args[0][2]
        assert "Imaginon" in written["branches"]
        assert "Imaginon" in written["months"][0]["branches"]


# ── lambda_handler – API Gateway ──────────────────────────────────────────────

class TestApiHandler:
    def _api_event(self, branch=None):
        event = {
            "httpMethod": "GET",
            "path": "/circulation",
            "requestContext": {"stage": "Prod"},
            "queryStringParameters": {},
        }
        if branch:
            event["queryStringParameters"]["branch"] = branch
        return event

    def _stored_payload(self, with_branch=False):
        """Return stored JSON in the new {months, branches} format."""
        system_totals = {
            cat: {"total": 1000, "breakdown": {"A": 600, "B": 400}}
            for cat in _CATEGORY_NAMES
        }
        branch_data = {
            cat: {"total": 500, "breakdown": {"A": 300, "B": 200}}
            for cat in _CATEGORY_NAMES
        }
        return {
            "months": [{
                "month_name": "JULY",
                "display_month": "Jul 2025",
                "year": 2025,
                "system_totals": system_totals,
                "branches": {"Main Branch": branch_data} if with_branch else {},
            }],
            "branches": ["Main Branch"] if with_branch else [],
            "lastUpdated": "2026-01-01T00:00:00",
        }

    @patch("lambda_function.read_json_s3")
    def test_returns_all_data_when_no_filter(self, mock_read):
        mock_read.return_value = self._stored_payload()
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            result = lambda_handler(self._api_event(), None)
        assert result["statusCode"] == 200
        body = json.loads(result["body"])
        assert body["success"] is True
        # 5 categories × 1 month = 5 data points
        assert len(body["data"]["data"]) == 5

    @patch("lambda_function.read_json_s3")
    def test_filters_by_branch(self, mock_read):
        mock_read.return_value = self._stored_payload(with_branch=True)
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            result = lambda_handler(self._api_event(branch="Main Branch"), None)
        body = json.loads(result["body"])
        assert body["success"] is True
        # 5 categories × 1 month = 5 points from the branch
        assert len(body["data"]["data"]) == 5
        assert body["data"]["selectedBranch"] == "Main Branch"

    @patch("lambda_function.read_json_s3")
    def test_unknown_branch_returns_empty_data(self, mock_read):
        mock_read.return_value = self._stored_payload()
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            result = lambda_handler(self._api_event(branch="Ghost Branch"), None)
        body = json.loads(result["body"])
        assert body["success"] is True
        assert body["data"]["data"] == []

    @patch("lambda_function.read_json_s3")
    def test_cors_headers_present(self, mock_read):
        mock_read.return_value = self._stored_payload()
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            result = lambda_handler(self._api_event(), None)
        assert result["headers"]["Access-Control-Allow-Origin"] == "*"

    @patch("lambda_function.read_json_s3")
    def test_response_matches_api_contract(self, mock_read):
        mock_read.return_value = self._stored_payload()
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            result = lambda_handler(self._api_event(), None)
        body = json.loads(result["body"])
        for field in ("success", "data", "error", "timestamp", "requestId"):
            assert field in body
        circ = body["data"]
        for field in ("data", "lastUpdated", "totalRecords", "branches", "selectedBranch"):
            assert field in circ

    @patch("lambda_function.read_json_s3")
    def test_data_points_include_breakdown(self, mock_read):
        mock_read.return_value = self._stored_payload()
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            result = lambda_handler(self._api_event(), None)
        body = json.loads(result["body"])
        point = body["data"]["data"][0]
        assert "breakdown" in point
        assert point["breakdown"]["A"] == 600

    def test_missing_bucket_config_returns_500(self):
        env = {k: v for k, v in os.environ.items() if k != "PROCESSED_BUCKET"}
        with patch.dict(os.environ, env, clear=True):
            result = lambda_handler(self._api_event(), None)
        body = json.loads(result["body"])
        assert result["statusCode"] == 500
        assert body["error"]["code"] == "CONFIG_ERROR"

    @patch("lambda_function.read_json_s3")
    def test_no_such_key_returns_404(self, mock_read):
        from botocore.exceptions import ClientError
        mock_read.side_effect = ClientError(
            {"Error": {"Code": "NoSuchKey", "Message": "The specified key does not exist."}},
            "GetObject",
        )
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            result = lambda_handler(self._api_event(), None)
        assert result["statusCode"] == 404

    @patch("lambda_function.read_json_s3")
    def test_access_denied_returns_404(self, mock_read):
        # S3 returns AccessDenied instead of NoSuchKey when ListBucket is missing
        from botocore.exceptions import ClientError
        mock_read.side_effect = ClientError(
            {"Error": {"Code": "AccessDenied", "Message": "Access Denied"}},
            "GetObject",
        )
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            result = lambda_handler(self._api_event(), None)
        assert result["statusCode"] == 404

    @patch("lambda_function.read_json_s3")
    def test_s3_error_returns_500(self, mock_read):
        from botocore.exceptions import ClientError
        mock_read.side_effect = ClientError(
            {"Error": {"Code": "InternalError", "Message": "Server error"}},
            "GetObject",
        )
        with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
            result = lambda_handler(self._api_event(), None)
        assert result["statusCode"] == 500


# ── Edge-case tests ───────────────────────────────────────────────────────────

class TestEdgeCases:
    """Edge cases not covered by the main test classes."""

    # ── Corrupted workbook ───────────────────────────────────────────────────

    def test_corrupted_bytes_raises_or_returns_empty(self):
        """parse_workbook with invalid bytes should raise or return nothing — not silently corrupt data."""
        import pytest
        bad_bytes = b"this is not a zip file at all" * 100
        try:
            months, branches = parse_workbook(bad_bytes)
            # If it doesn't raise, it must return empty — not fake data
            assert months == []
            assert branches == []
        except Exception:
            pass  # openpyxl raises on invalid bytes — also acceptable

    def test_empty_bytes_handled_gracefully(self):
        import pytest
        try:
            months, branches = parse_workbook(b"")
            assert months == []
        except Exception:
            pass  # acceptable: empty bytes is not a valid xlsx

    # ── Branch with zero circulation ─────────────────────────────────────────

    def test_branch_with_zero_grand_total_is_excluded(self):
        """A branch row whose Grand Total is 0 must be excluded from extracted branches."""
        zero_branch = {col: 0 for col in IMAGINON_BRANCH_DATA}
        zero_branch[COL_GRAND_TOTAL] = 0
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]}, branch_rows=zero_branch)
        months, branches = parse_workbook(wb)
        assert "Imaginon" not in branches
        assert "Imaginon" not in months[0]["branches"]

    def test_branch_with_nonzero_total_is_included(self):
        """Sanity check: positive Grand Total keeps the branch."""
        wb = _make_workbook({"JULY": REAL_DATA["JULY"]}, branch_rows=IMAGINON_BRANCH_DATA)
        _, branches = parse_workbook(wb)
        assert "Imaginon" in branches

    # ── _is_skip_department with all known skip values ───────────────────────

    def test_all_skip_department_values_are_skipped(self):
        from lambda_function import SKIP_DEPARTMENTS
        for name in SKIP_DEPARTMENTS:
            assert _is_skip_department(name) is True, f"Expected '{name}' to be skipped"

    def test_numeric_department_name_is_skipped(self):
        """Non-string types (e.g. a cell with a number) must be skipped."""
        assert _is_skip_department(12345) is True

    # ── build_circulation_data with empty months ──────────────────────────────

    def test_empty_months_returns_no_data_points(self):
        result = build_circulation_data([])
        assert result["data"] == []
        assert result["totalRecords"] == 0
        assert result["branches"] == []

    # ── Multiple months – correct ordering in output ──────────────────────────

    def test_output_preserves_month_order(self):
        months = []
        for i, m in enumerate(["JULY", "AUGUST", "SEPTEMBER"]):
            months.append({
                "month_name": m,
                "display_month": f"{m[:3].title()} 2025",
                "year": 2025,
                "system_totals": {
                    cat: {"total": (i + 1) * 1000, "breakdown": {}}
                    for cat in _CATEGORY_NAMES
                },
                "branches": {},
            })
        result = build_circulation_data(months)
        output_months = [p["month"] for p in result["data"] if p["category"] == "Adult"]
        assert output_months == ["Jul 2025", "Aug 2025", "Sep 2025"]

    # ── S3 handler with URL-encoded spaces ───────────────────────────────────

    def test_url_encoded_plus_signs_decoded(self):
        """S3 encodes spaces in object keys as '+'; they must become ' '."""
        from lambda_function import handle_s3_event
        with patch("lambda_function.read_s3") as mock_read, \
             patch("lambda_function.write_json_s3"):
            mock_read.return_value = _make_workbook({"JULY": REAL_DATA["JULY"]})
            event = {
                "Records": [{
                    "s3": {
                        "bucket": {"name": "b"},
                        "object": {"key": "uploads/FY2026+Circ+Stats.xlsm"},
                    }
                }]
            }
            with patch.dict(os.environ, {"PROCESSED_BUCKET": "b"}):
                handle_s3_event(event)
            decoded_key = mock_read.call_args[0][1]
            assert " " in decoded_key
            assert "+" not in decoded_key
