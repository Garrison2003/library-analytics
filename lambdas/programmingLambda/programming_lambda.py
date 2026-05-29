"""
Library Analytics – Programming Statistics Lambda

Parses programming statistics .xlsx files uploaded to S3 and serves processed
graph data to the frontend dashboard via API Gateway.

Features:
  - Extracts monthly in-person attendance and program count data
  - Identifies department/branch by 3-letter code in filename (e.g., IMG)
  - Returns data organized by month in fiscal year format
  - Supports branch-level filtering via query parameter

Trigger:  S3 PUT on uploads/programming/*.xlsx
API:      GET /programming?branch={CODE}

Response format:
  {
    "success": true,
    "data": {
      "branch": "IMG",
      "branchName": "Imaginon",
      "data": [
        { "month": "07/24", "attendance": 23237, "programs": 167, "date": "2024-07-01" },
        ...
      ],
      "months": ["07/24", "08/24", ...],
      "attendance": [23237, ...],
      "programs": [167, ...],
      "lastUpdated": "2026-05-28T...",
      "dataFound": true
    }
  }
"""

import json
import logging
import os
import re
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import unquote_plus

import boto3
import openpyxl
from botocore.exceptions import ClientError

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client("s3")

# ── Constants ────────────────────────────────────────────────────────────────

# Branch code to full name mapping
BRANCH_CODE_MAP = {
    "IMG": "Imaginon",
    "MAI": "Main",
    "PLZ": "Plaza Midwood",
    "NOR": "Northlake",
    "CHS": "Charlotte",
    "SPA": "Spangler",
    "CAR": "Carmel",
    "COM": "Community",
    "EAS": "East",
    "WES": "West",
}

# Month display format: MM/YY based on fiscal year (July-June)
FISCAL_MONTHS = [
    ("07", "24"), ("08", "24"), ("09", "24"), ("10", "24"), ("11", "24"), ("12", "24"),
    ("01", "25"), ("02", "25"), ("03", "25"), ("04", "25"), ("05", "25"), ("06", "25"),
    ("07", "25"), ("08", "25"), ("09", "25"), ("10", "25"), ("11", "25"), ("12", "25"),
]

PROCESSED_DIR = "processed/programming"

# ── Fiscal-year filtering ─────────────────────────────────────────────────────

def _two_fy_cutoff(now: Optional[datetime] = None) -> datetime:
    """
    Return the July 1 start date of the fiscal year that began two FYs ago.
    For example, on any date in FY26 (Jul 2025 – Jun 2026) the cutoff is 2024-07-01,
    so the returned data spans FY25 and FY26.
    """
    if now is None:
        now = datetime.now(timezone.utc)
    fy_start_year = now.year if now.month >= 7 else now.year - 1
    return datetime(fy_start_year - 1, 7, 1)


def _filter_to_last_two_fiscal_years(payload: dict, now: Optional[datetime] = None) -> dict:
    """Return a copy of payload with data filtered to the last two fiscal years."""
    cutoff = _two_fy_cutoff(now)
    filtered = [
        r for r in payload.get("data", [])
        if datetime.fromisoformat(r["date"]) >= cutoff
    ]
    return {
        **payload,
        "data": filtered,
        "months": [r["month"] for r in filtered],
        "attendance": [r["attendance"] for r in filtered],
        "programs": [r["programs"] for r in filtered],
    }


# ── Helpers ──────────────────────────────────────────────────────────────────

def _env(name: str, fallback: str = "") -> str:
    """Get environment variable or fallback."""
    return os.environ.get(name, fallback)


def _safe_int(val: Any) -> int:
    """Safely convert value to int, returning 0 on failure."""
    if val is None:
        return 0
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _extract_branch_code(filename: str) -> Optional[str]:
    """
    Extract 3-letter branch code from filename.
    Example: "IMG Monthly Stats FY26-28.xlsx" → "IMG"
    """
    match = re.match(r"^([A-Z]{3})\s+", filename)
    return match.group(1) if match else None


def _get_branch_name(code: str) -> str:
    """Get full branch name from code, or return code if not found."""
    return BRANCH_CODE_MAP.get(code, code)


# ── Excel Parsing ────────────────────────────────────────────────────────────

def _find_data_sheet(workbook) -> Optional[str]:
    """
    Find the appropriate data sheet in the workbook.
    Looks for "Program 23-26", "Program 23-27", or similar patterns.
    Falls back to "Teen Programs" or "Juv Programs".
    """
    for sheet_name in workbook.sheetnames:
        if "Program 23" in sheet_name or "Program 24" in sheet_name:
            return sheet_name
    for sheet_name in ["Teen Programs", "Juv Programs"]:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    return None


def _parse_date_row(row: List) -> Optional[Tuple[int, int]]:
    """
    Parse the first column (date) from a data row.
    Returns (month_number, year) or None if invalid.
    Example: datetime(2024, 7, 1) → (7, 24)
    """
    date_cell = row[0] if row else None
    if not date_cell:
        return None

    try:
        if hasattr(date_cell, "month") and hasattr(date_cell, "year"):
            month = date_cell.month
            year = date_cell.year
            return (month, year % 100)
    except (AttributeError, TypeError):
        pass

    return None


def _month_year_to_display(month: int, year: int) -> str:
    """Convert month (1-12) and year (24, 25...) to MM/YY format."""
    return f"{month:02d}/{year:02d}"


def parse_workbook(file_bytes: bytes) -> Dict[str, Any]:
    """
    Parse programming .xlsx and extract monthly data.

    Returns dict with:
      data: List of monthly records with attendance & program counts
      months: List of MM/YY labels
      attendance: List of attendance values
      programs: List of program counts
    """
    file_obj = BytesIO(file_bytes)
    wb = openpyxl.load_workbook(file_obj, data_only=True)

    data_sheet_name = _find_data_sheet(wb)
    if not data_sheet_name:
        logger.warning("No data sheet found in workbook. Available: %s", wb.sheetnames)
        wb.close()
        return {
            "data": [],
            "months": [],
            "attendance": [],
            "programs": [],
        }

    ws = wb[data_sheet_name]
    logger.info("Parsing sheet: %s", data_sheet_name)

    result_data = []
    months_list = []
    attendance_list = []
    programs_list = []

    # Iterate through rows starting from row 2 (row 1 is headers)
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            # Skip header row
            continue

        if not row or not row[0]:
            # Skip empty rows
            continue

        # Parse date from first column
        date_info = _parse_date_row(row)
        if not date_info:
            continue

        month, year = date_info

        # Column mapping (0-based after date column):
        # Col 0: Date
        # Col 1: TOTAL ATTENDANCE IN-PERSON
        # Col 2: TOTAL PROGRAMS
        # (Further columns: TOTAL VIRTUAL ATTENDANCE, increases, etc.)

        attendance = _safe_int(row[1]) if len(row) > 1 else 0
        programs = _safe_int(row[2]) if len(row) > 2 else 0

        # Skip rows with no meaningful data
        if attendance == 0 and programs == 0:
            continue

        month_display = _month_year_to_display(month, year)
        months_list.append(month_display)
        attendance_list.append(attendance)
        programs_list.append(programs)

        result_data.append({
            "month": month_display,
            "attendance": attendance,
            "programs": programs,
            "date": f"20{year:02d}-{month:02d}-01",
        })

    wb.close()

    logger.info("Parsed %d months of data", len(result_data))

    return {
        "data": result_data,
        "months": months_list,
        "attendance": attendance_list,
        "programs": programs_list,
    }


# ── S3 I/O ───────────────────────────────────────────────────────────────────

def read_s3(bucket: str, key: str) -> bytes:
    """Read file from S3."""
    return s3.get_object(Bucket=bucket, Key=key)["Body"].read()


def write_json_s3(bucket: str, key: str, payload: dict) -> None:
    """Write JSON payload to S3."""
    s3.put_object(
        Bucket=bucket,
        Key=key,
        Body=json.dumps(payload, indent=2),
        ContentType="application/json",
    )
    logger.info("Wrote s3://%s/%s", bucket, key)


def read_json_s3(bucket: str, key: str) -> dict:
    """Read JSON from S3."""
    return json.loads(s3.get_object(Bucket=bucket, Key=key)["Body"].read())


# ── API Response Helpers ─────────────────────────────────────────────────────

def _api_ok(data: Any) -> dict:
    """Successful API response."""
    return _api_response(200, {
        "success": True,
        "data": data,
        "error": None,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "requestId": f"req_{uuid.uuid4().hex[:12]}",
    })


def _api_err(status: int, code: str, message: str) -> dict:
    """Error API response."""
    return _api_response(status, {
        "success": False,
        "data": None,
        "error": {"code": code, "message": message},
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "requestId": f"req_{uuid.uuid4().hex[:12]}",
    })


def _api_response(status: int, body: dict) -> dict:
    """Format API response with CORS headers."""
    return {
        "statusCode": status,
        "headers": {
            "Content-Type": "application/json",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "Content-Type, Authorization",
        },
        "body": json.dumps(body),
    }


# ── Handlers ─────────────────────────────────────────────────────────────────

def handle_s3_event(event: dict) -> dict:
    """S3 PUT trigger. Downloads .xlsx, parses it, writes processed JSON per branch."""
    try:
        record = event["Records"][0]["s3"]
        source_bucket = record["bucket"]["name"]
        source_key = unquote_plus(record["object"]["key"])

        logger.info("Processing s3://%s/%s", source_bucket, source_key)

        # Extract branch code from filename
        filename = source_key.split("/")[-1]
        branch_code = _extract_branch_code(filename)

        if not branch_code:
            logger.error("Cannot extract branch code from filename: %s", filename)
            return {
                "statusCode": 400,
                "body": json.dumps({"error": "Invalid filename format"}),
            }

        # Read and parse file
        file_bytes = read_s3(source_bucket, source_key)
        parsed = parse_workbook(file_bytes)

        if not parsed["data"]:
            logger.warning("No valid data found in file")
            return {
                "statusCode": 400,
                "body": json.dumps({"error": "No valid data in workbook"}),
            }

        # Prepare output JSON
        output_payload = {
            "branch": branch_code,
            "branchName": _get_branch_name(branch_code),
            "data": parsed["data"],
            "months": parsed["months"],
            "attendance": parsed["attendance"],
            "programs": parsed["programs"],
            "lastUpdated": datetime.now(timezone.utc).isoformat(),
            "dataFound": True,
        }

        # Write to S3: processed/programming/{BRANCH_CODE}.json
        dest_bucket = _env("PROCESSED_BUCKET", source_bucket)
        dest_key = f"{PROCESSED_DIR}/{branch_code}.json"
        write_json_s3(dest_bucket, dest_key, output_payload)

        logger.info("Successfully processed %s for branch %s", filename, branch_code)
        return {
            "statusCode": 200,
            "body": json.dumps({
                "message": "Programming data processed",
                "branch": branch_code,
                "dataPoints": len(parsed["data"]),
            }),
        }

    except Exception as exc:
        logger.exception("Error processing S3 event")
        return {
            "statusCode": 500,
            "body": json.dumps({"error": str(exc)}),
        }


def handle_api_request(event: dict) -> dict:
    """GET /programming?branch={CODE} - Return programming data for a branch."""
    try:
        bucket = _env("PROCESSED_BUCKET")
        if not bucket:
            return _api_err(500, "CONFIG_ERROR", "PROCESSED_BUCKET not set")

        # Get branch code from query parameters
        query_params = event.get("queryStringParameters") or {}
        branch_code = query_params.get("branch", "").upper()

        if not branch_code:
            return _api_err(400, "MISSING_PARAMETER", "branch query parameter required")

        # Try to read branch-specific file
        key = f"{PROCESSED_DIR}/{branch_code}.json"

        try:
            payload = read_json_s3(bucket, key)
            payload = _filter_to_last_two_fiscal_years(payload)
            return _api_ok(payload)
        except ClientError as exc:
            code = exc.response["Error"]["Code"]
            if code in ("NoSuchKey", "404"):
                # Return no-data response instead of 404
                return _api_ok({
                    "branch": branch_code,
                    "branchName": _get_branch_name(branch_code),
                    "data": [],
                    "months": [],
                    "attendance": [],
                    "programs": [],
                    "dataFound": False,
                    "lastUpdated": datetime.now(timezone.utc).isoformat(),
                })
            raise

    except Exception as exc:
        logger.exception("Error handling API request")
        return _api_err(500, "INTERNAL_ERROR", str(exc))


# ── Lambda Handler Entrypoint ────────────────────────────────────────────────

def lambda_handler(event, context):
    """Main Lambda entry point. Routes S3 events vs API Gateway requests."""
    logger.info("Event: %s", json.dumps(event, default=str)[:500])

    # Determine event type
    if "Records" in event:
        # S3 event
        return handle_s3_event(event)
    else:
        # API Gateway request
        return handle_api_request(event)
